import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


def export_to_excel(rows: list, output_dir: str = ".") -> str:
    """
    Write inventory rows to a formatted .xlsx file.
    Returns the absolute path of the created file.
    Raises RuntimeError if openpyxl is not installed.
    """
    if not OPENPYXL_OK:
        raise RuntimeError(
            "openpyxl is not installed. Run: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    PRIMARY_COLOR = "D4430A"
    HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HEADER_FILL   = PatternFill("solid", fgColor=PRIMARY_COLOR)
    HEADER_ALIGN  = Alignment(horizontal="center", vertical="center")
    DATA_FONT     = Font(name="Calibri", size=10)
    EVEN_FILL     = PatternFill("solid", fgColor="FFF3EC")
    TOTAL_FONT    = Font(name="Calibri", bold=True, size=10, color=PRIMARY_COLOR)
    thin          = Side(style="thin", color="E0C4B8")
    CELL_BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)

    #Title row
    ws.merge_cells("A1:G1")
    tc = ws["A1"]
    tc.value     = "Martus Sari-Sari Store — Product Inventory"
    tc.font      = Font(name="Calibri", bold=True, size=14, color=PRIMARY_COLOR)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    sc = ws["A2"]
    sc.value     = f"Exported: {datetime.now().strftime('%B %d, %Y  %I:%M %p')}"
    sc.font      = Font(name="Calibri", italic=True, size=9, color="7A5C4F")
    sc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    #Header row
    HEADERS    = ["ID", "Product Code", "Product Name", "Category",
                  "Price (P)", "Stock (pcs)", "Added By"]
    COL_WIDTHS = [6, 14, 28, 18, 14, 13, 14]

    for col, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell            = ws.cell(row=3, column=col, value=hdr)
        cell.font       = HEADER_FONT
        cell.fill       = HEADER_FILL
        cell.alignment  = HEADER_ALIGN
        cell.border     = CELL_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 20

    #Data rows
    data_start = 4
    for row_idx, r in enumerate(rows, start=data_start):
        is_even = (row_idx - data_start) % 2 == 1
        values  = [r[0], r[1], r[2], r[3], r[4], r[5],
                   r[6] if len(r) > 6 else "—"]
        for col, val in enumerate(values, start=1):
            cell            = ws.cell(row=row_idx, column=col, value=val)
            cell.font       = DATA_FONT
            cell.border     = CELL_BORDER
            cell.alignment  = Alignment(
                horizontal="center" if col != 3 else "left",
                vertical="center",
            )
            if is_even:
                cell.fill = EVEN_FILL
            if col == 5:
                cell.number_format = "#,##0.00"

    #Totals row
    last_data = data_start + len(rows) - 1
    total_row = last_data + 1

    lbl           = ws.cell(row=total_row, column=4, value="TOTAL")
    lbl.font      = TOTAL_FONT
    lbl.alignment = Alignment(horizontal="right")

    ps            = ws.cell(row=total_row, column=5,
                            value=f"=SUM(E{data_start}:E{last_data})")
    ps.font           = TOTAL_FONT
    ps.number_format  = "#,##0.00"
    ps.alignment      = Alignment(horizontal="center")

    ss            = ws.cell(row=total_row, column=6,
                            value=f"=SUM(F{data_start}:F{last_data})")
    ss.font       = TOTAL_FONT
    ss.alignment  = Alignment(horizontal="center")

    ws.freeze_panes = "A4"

    filename  = f"inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    full_path = os.path.abspath(os.path.join(output_dir, filename))
    wb.save(full_path)
    return full_path
